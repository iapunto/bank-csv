# -*- coding: utf-8 -*-
"""
Fichero: excel_writer.py
Proyecto: Extractor de Movimientos Bancarios con IA

Desarrollado por: IA Punto Soluciones Tecnológicas
Para: Industrias Pico
Responsable: MEng Sergio Rondón
Fecha de Creación: 08/09/2025

Descripción:
Este módulo proporciona la funcionalidad para escribir la lista de transacciones
extraídas en un archivo Excel (.xlsx) con un formato estandarizado.
"""

import logging
from typing import List
import openpyxl
from openpyxl.styles import Font, Alignment
from .data_models import Transaccion

# Configurar logging
logger = logging.getLogger(__name__)


def escribir_transacciones_a_excel(transacciones: List[Transaccion], output_path: str) -> bool:
    """
    Escribe una lista de objetos Transaccion en un archivo Excel (.xlsx).

    Args:
        transacciones: Una lista que contiene los objetos Transaccion extraídos.
        output_path: La ruta completa del archivo donde se guardará el Excel.

    Returns:
        True si el archivo se escribió correctamente, False si ocurrió un error.
    """
    if not transacciones:
        logger.warning("No se encontraron transacciones para escribir en el Excel.")
        return False

    if not all(isinstance(t, Transaccion) for t in transacciones):
        logger.error("La lista contiene elementos que no son instancias de Transaccion")
        return False

    # Definimos las cabeceras
    headers = ['Día', 'Etiqueta', 'Debit', 'Credit']
    
    logger.info(f"Escribiendo {len(transacciones)} transacciones en el archivo: {output_path}")

    try:
        # Crear un nuevo libro de trabajo y seleccionar la hoja activa
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Movimientos"

        # Estilo para las cabeceras
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Escribir las cabeceras con estilo
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Escribir cada transacción en una nueva fila
        for i, transaccion in enumerate(transacciones, start=2):
            try:
                row_data = transaccion.model_dump()
                
                sheet.cell(row=i, column=1, value=row_data['fecha'])
                sheet.cell(row=i, column=2, value=row_data['descripcion'])
                
                debito = row_data.get('debito')
                if debito is not None:
                    debit_cell = sheet.cell(row=i, column=3, value=float(debito))
                    debit_cell.number_format = '#,##0.00'

                credito = row_data.get('credito')
                if credito is not None:
                    credit_cell = sheet.cell(row=i, column=4, value=float(credito))
                    credit_cell.number_format = '#,##0.00'

            except Exception as e:
                logger.error(f"Error al escribir la transacción {i-1}: {e}")
                continue
        
        # Ajustar el ancho de las columnas
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # Guardar el libro de trabajo
        workbook.save(output_path)
        logger.info("Archivo Excel generado exitosamente.")
        return True

    except Exception as e:
        logger.error(f"Ocurrió un error inesperado al escribir el Excel: {e}")
        return False
